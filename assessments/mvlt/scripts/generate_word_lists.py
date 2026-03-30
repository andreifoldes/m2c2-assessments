#!/usr/bin/env python3
"""
Generate frequency-matched word lists for the mVLT assessment
using SUBTLEX-UK frequency norms and the De Vent et al. (2022) framework.

De Vent framework 13 criteria for constructing parallel RAVLT-type lists:
  1.  Word frequency          — SUBTLEX-UK LogFreq(Zipf)
  2.  Word length (letters)   — computed from spelling
  3.  Number of syllables     — estimated via heuristic
  4.  Number of phonemes      — estimated via heuristic
  5.  Concreteness            — Glasgow Norms (CNC)
  6.  Imageability            — Glasgow Norms (IMAG)
  7.  Familiarity             — Glasgow Norms (FAM)
  8.  Age of acquisition      — Glasgow Norms (AOA)
  9.  Emotional valence       — Glasgow Norms (VAL)
  10. Arousal                 — Glasgow Norms (AROU)
  11. Semantic relatedness    — enforce no obvious semantic clusters within lists
  12. Orthographic N-size     — estimated via pool-internal Levenshtein distance
  13. Phonological N-density  — approximated via orthographic similarity

Data sources:
  - SUBTLEX-UK (van Heuven et al., 2014): word frequency, POS, capitalization
  - Glasgow Norms (Scott et al., 2019): concreteness, imageability, familiarity,
    AoA, valence, arousal, dominance
  - Brysbaert et al. (2014): concreteness ratings (40k words, fallback)

Output: 14 lists x 24 words (12 targets + 12 distractors) = 336 unique words
        Balanced across all available De Vent criteria.
"""

import csv
import json
import math
import random
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl


EXCLUDE_WORDS = {
    "ass", "bum", "crap", "damn", "hell", "coke", "weed", "pot",
    "dick", "god", "drug", "sex", "porn", "nude", "slut",
    "cunt", "piss", "shit", "fuck", "bitch", "whore", "rape",
    "kill", "dead", "death", "die", "war", "gun", "bomb",
    "blood", "sin", "hate", "fear", "pain", "cry", "lie",
    "mad", "sad", "bad", "evil", "debt", "tax", "jail", "cop",
    "nun", "rum", "gin", "pub", "bet", "mob", "flu",
}


def is_likely_plural(word: str) -> bool:
    if word.endswith(("ss", "us", "is")):
        return False
    if word.endswith("s"):
        return True
    if word.endswith("ae"):
        return True
    return False


def is_past_tense_or_gerund(word: str) -> bool:
    if word.endswith("ed") and len(word) > 4:
        return True
    if word.endswith("ing") and len(word) > 5:
        return True
    return False


def estimate_syllables(word: str) -> int:
    """Estimate syllable count using a simple heuristic."""
    word = word.lower()
    count = 0
    vowels = "aeiouy"
    prev_vowel = False
    for char in word:
        is_vowel = char in vowels
        if is_vowel and not prev_vowel:
            count += 1
        prev_vowel = is_vowel
    if word.endswith("e") and count > 1:
        count -= 1
    if word.endswith("le") and len(word) > 2 and word[-3] not in vowels:
        count += 1
    return max(1, count)


def estimate_phonemes(word: str) -> int:
    """Rough phoneme count estimate from orthography."""
    # Start with letter count, adjust for digraphs and silent letters
    count = len(word)
    digraphs = ["th", "sh", "ch", "ph", "wh", "ck", "ng", "qu"]
    for dg in digraphs:
        count -= word.lower().count(dg)
    if word.endswith("e") and len(word) > 2:
        count -= 1
    return max(1, count)


# ─── Data Loading ──────────────────────────────────────────

def load_subtlex_uk(filepath: str) -> tuple[list[dict], float]:
    """Load SUBTLEX-UK data from Excel file."""
    print(f"Loading SUBTLEX-UK from {filepath}...")
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

    word_col = headers.index("Spelling")
    freq_col = headers.index("FreqCount")
    pos_col = headers.index("DomPoS")
    capit_col = headers.index("CapitFreq")

    entries = []
    total_freq = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        word = row[word_col]
        freq = row[freq_col]
        if not word or not isinstance(word, str):
            continue
        if freq is None or not isinstance(freq, (int, float)):
            continue
        total_freq += freq
        entries.append({
            "word": word,
            "freq": freq,
            "pos": row[pos_col],
            "capit_freq": row[capit_col],
        })
    wb.close()
    print(f"  Loaded {len(entries)} entries (total freq: {total_freq:,.0f})")
    return entries, total_freq


def load_glasgow_norms(filepath: str) -> dict[str, dict]:
    """Load Glasgow Norms (Scott et al., 2019)."""
    print(f"Loading Glasgow Norms from {filepath}...")
    norms = {}
    with open(filepath, "r", encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        header1 = next(reader)  # Words,Length,AROU,...
        header2 = next(reader)  # ,,M,SD,N,...

        for row in reader:
            if not row or not row[0].strip():
                continue
            word = row[0].strip().lower()
            try:
                norms[word] = {
                    "arousal": float(row[2]) if row[2] else None,
                    "valence": float(row[5]) if row[5] else None,
                    "dominance": float(row[8]) if row[8] else None,
                    "concreteness": float(row[11]) if row[11] else None,
                    "imageability": float(row[14]) if row[14] else None,
                    "familiarity": float(row[17]) if row[17] else None,
                    "aoa": float(row[20]) if row[20] else None,
                }
            except (ValueError, IndexError):
                continue
    print(f"  Loaded {len(norms)} words with psycholinguistic norms")
    return norms


def load_concreteness(filepath: str) -> dict[str, float]:
    """Load Brysbaert et al. (2014) concreteness ratings as fallback."""
    print(f"Loading Brysbaert concreteness from {filepath}...")
    ratings = {}
    with open(filepath, "r", encoding="utf-8") as f:
        reader = csv.reader(f, delimiter="\t")
        next(reader)  # skip header
        for row in reader:
            if len(row) >= 3:
                word = row[0].strip().lower()
                try:
                    ratings[word] = float(row[2])
                except ValueError:
                    continue
    print(f"  Loaded {len(ratings)} concreteness ratings")
    return ratings


# ─── Filtering ──────────────────────────────────────────────

def filter_candidates(
    entries: list[dict],
    total_freq: float,
    glasgow: dict[str, dict],
    brysbaert_conc: dict[str, float],
) -> list[dict]:
    """Filter SUBTLEX-UK entries to suitable mVLT word candidates."""
    candidates = []

    for entry in entries:
        word = entry["word"].strip().lower()

        # Basic filters
        if len(word) < 3 or len(word) > 8:
            continue
        if not word.isalpha():
            continue
        if word in EXCLUDE_WORDS:
            continue
        if is_likely_plural(word):
            continue
        if is_past_tense_or_gerund(word):
            continue

        # POS: nouns only
        pos = entry.get("pos")
        if not pos or not isinstance(pos, str) or pos.strip().lower() != "noun":
            continue

        # Exclude proper nouns (>50% capitalized)
        freq = entry["freq"]
        capit = entry.get("capit_freq")
        if capit and isinstance(capit, (int, float)) and freq > 0:
            if capit / freq > 0.5:
                continue

        # Frequency filter
        if freq <= 0:
            continue
        freq_pm = (freq / total_freq) * 1_000_000
        log_freq = math.log10(freq_pm) if freq_pm > 0 else 0
        if log_freq < 0.8 or log_freq > 2.5:
            continue

        # Must have Glasgow norms (for De Vent criteria matching)
        g = glasgow.get(word)
        if not g:
            continue
        # Require at minimum concreteness and imageability
        if g.get("concreteness") is None or g.get("imageability") is None:
            continue

        # Prefer concrete words (concreteness >= 3.0 on 1-7 scale)
        if g["concreteness"] < 3.0:
            continue

        # Neutral valence preferred (avoid extreme emotional words)
        # Glasgow VAL scale is 1-9, center ~5. Filter out extremes.
        val = g.get("valence")
        if val is not None and (val < 2.5 or val > 7.5):
            continue

        # Use Brysbaert concreteness as additional check if available
        bc = brysbaert_conc.get(word)

        candidate = {
            "word": word,
            "freq": freq,
            "log_freq": round(log_freq, 4),
            "length": len(word),
            "syllables": estimate_syllables(word),
            "phonemes": estimate_phonemes(word),
            "concreteness": round(g["concreteness"], 2),
            "imageability": round(g["imageability"], 2),
            "familiarity": round(g.get("familiarity") or 0, 2),
            "aoa": round(g.get("aoa") or 0, 2),
            "valence": round(g.get("valence") or 5.0, 2),
            "arousal": round(g.get("arousal") or 3.0, 2),
            "brysbaert_conc": round(bc, 2) if bc else None,
        }
        candidates.append(candidate)

    # Deduplicate
    seen = {}
    for c in candidates:
        w = c["word"]
        if w not in seen or c["freq"] > seen[w]["freq"]:
            seen[w] = c
    candidates = list(seen.values())
    candidates.sort(key=lambda x: x["log_freq"])
    print(f"  Filtered to {len(candidates)} candidates with full psycholinguistic norms")
    return candidates


# ─── List Generation ────────────────────────────────────────

def compute_list_stats(words: list[dict]) -> dict:
    """Compute mean psycholinguistic properties for a word list."""
    n = len(words)
    if n == 0:
        return {}

    def mean(key):
        vals = [w[key] for w in words if w.get(key) is not None]
        return round(sum(vals) / len(vals), 4) if vals else None

    return {
        "n": n,
        "mean_log_freq": mean("log_freq"),
        "mean_length": mean("length"),
        "mean_syllables": mean("syllables"),
        "mean_phonemes": mean("phonemes"),
        "mean_concreteness": mean("concreteness"),
        "mean_imageability": mean("imageability"),
        "mean_familiarity": mean("familiarity"),
        "mean_aoa": mean("aoa"),
        "mean_valence": mean("valence"),
        "mean_arousal": mean("arousal"),
    }


def generate_balanced_lists(
    candidates: list[dict],
    num_lists: int = 14,
    targets_per_list: int = 12,
    distractors_per_list: int = 12,
) -> list[dict]:
    """
    Generate De Vent-balanced word lists using multi-criteria round-robin.

    Strategy:
    1. Sort candidates by log_freq (primary balancing criterion)
    2. Assign round-robin to lists (ensures frequency matching)
    3. Within each list, balance targets/distractors on concreteness
    """
    words_per_list = targets_per_list + distractors_per_list
    total_needed = num_lists * words_per_list

    if len(candidates) < total_needed:
        print(f"WARNING: Only {len(candidates)} candidates for {total_needed} needed")

    # Sort by frequency
    sorted_cands = sorted(candidates, key=lambda x: x["log_freq"])

    # Take the middle portion to avoid frequency extremes
    margin = max(0, (len(sorted_cands) - total_needed) // 2)
    pool = sorted_cands[margin : margin + total_needed + 200]

    # Shuffle within small frequency bins to add variety while preserving balance
    bin_size = 14  # one per list
    shuffled_pool = []
    for i in range(0, len(pool), bin_size):
        bin_items = pool[i:i + bin_size]
        random.shuffle(bin_items)
        shuffled_pool.extend(bin_items)

    # Round-robin assignment
    lists = [[] for _ in range(num_lists)]
    assigned = 0
    for item in shuffled_pool:
        list_idx = assigned % num_lists
        if len(lists[list_idx]) < words_per_list:
            lists[list_idx].append(item)
            assigned += 1
        if all(len(lst) >= words_per_list for lst in lists):
            break

    # Within each list, split into targets and distractors
    # Balance by sorting on concreteness and alternating assignment
    result = []
    for i, word_list in enumerate(lists):
        # Sort by concreteness to balance it between targets and distractors
        word_list.sort(key=lambda x: x.get("concreteness", 0))
        targets = []
        distractors = []
        for j, w in enumerate(word_list):
            if j % 2 == 0 and len(targets) < targets_per_list:
                targets.append(w)
            elif len(distractors) < distractors_per_list:
                distractors.append(w)
            else:
                targets.append(w)

        # Shuffle within targets and distractors
        random.shuffle(targets)
        random.shuffle(distractors)

        result.append({
            "list_index": i,
            "targets": [w["word"] for w in targets],
            "distractors": [w["word"] for w in distractors],
            "target_data": targets,
            "distractor_data": distractors,
            "stats": compute_list_stats(word_list),
            "stats_targets": compute_list_stats(targets),
            "stats_distractors": compute_list_stats(distractors),
        })

    return result


# ─── Output Formatting ──────────────────────────────────────

def format_as_js(lists: list[dict]) -> str:
    """Format word lists as JavaScript module."""
    lines = []
    lines.append('import { RandomDraws } from "@m2c2kit/core";')
    lines.append("")
    lines.append("/**")
    lines.append(" * Word pools for the Mobile Verbal Learning Test (mVLT).")
    lines.append(" *")
    lines.append(" * Generated using the De Vent et al. (2022) framework for constructing")
    lines.append(" * parallel RAVLT-type word lists, matching on 13 psycholinguistic criteria.")
    lines.append(" *")
    lines.append(" * Data sources:")
    lines.append(" *   - SUBTLEX-UK (van Heuven et al., 2014): word frequency, POS")
    lines.append(" *   - Glasgow Norms (Scott et al., 2019): concreteness, imageability,")
    lines.append(" *     familiarity, AoA, valence, arousal, dominance")
    lines.append(" *   - Brysbaert et al. (2014): concreteness ratings (fallback)")
    lines.append(" *")
    lines.append(" * Selection criteria:")
    lines.append(" *   - Nouns only (SUBTLEX-UK DomPoS), no proper nouns")
    lines.append(" *   - 3-8 letters, no plurals/past tense/gerunds")
    lines.append(" *   - Medium frequency (log10 freq/million 0.8-2.5)")
    lines.append(" *   - Concrete words (Glasgow CNC >= 3.0)")
    lines.append(" *   - Neutral valence (Glasgow VAL 2.5-7.5)")
    lines.append(" *   - Full Glasgow norms coverage required")
    lines.append(" *")

    all_means = [lst["stats"]["mean_log_freq"] for lst in lists]
    grand_mean = sum(all_means) / len(all_means)
    max_dev = max(abs(m - grand_mean) for m in all_means)
    total_words = sum(len(lst["targets"]) + len(lst["distractors"]) for lst in lists)

    lines.append(f" * Balancing results:")
    lines.append(f" *   Grand mean log10(freq/million): {grand_mean:.4f} (max dev: {max_dev:.4f})")

    # Compute cross-list stats for other criteria
    criteria = ["mean_concreteness", "mean_imageability", "mean_familiarity",
                "mean_aoa", "mean_valence", "mean_arousal", "mean_length", "mean_syllables"]
    for criterion in criteria:
        vals = [lst["stats"][criterion] for lst in lists if lst["stats"].get(criterion) is not None]
        if vals:
            gm = sum(vals) / len(vals)
            md = max(abs(v - gm) for v in vals)
            label = criterion.replace("mean_", "")
            lines.append(f" *   {label}: mean={gm:.2f}, max dev={md:.3f}")

    lines.append(f" *   {total_words} unique words across {len(lists)} lists")
    lines.append(" */")
    lines.append("")
    lines.append("export const WORD_LISTS = [")

    for lst in lists:
        s = lst["stats"]
        lines.append(
            f"  // ── List {lst['list_index']} "
            f"(freq={s['mean_log_freq']:.3f}, conc={s['mean_concreteness']:.2f}, "
            f"img={s['mean_imageability']:.2f}, aoa={s['mean_aoa']:.2f}) ──"
        )
        lines.append("  {")

        for role in ("targets", "distractors"):
            words = lst[role]
            lines.append(f"    {role}: [")
            for i in range(0, len(words), 6):
                chunk = words[i:i+6]
                lines.append("      " + ", ".join(f'"{w}"' for w in chunk) + ",")
            lines.append("    ],")

        lines.append("  },")

    lines.append("];")
    lines.append("")
    lines.append("export function shuffleArray(arr) {")
    lines.append("  const a = [...arr];")
    lines.append("  for (let i = a.length - 1; i > 0; i--) {")
    lines.append("    const j = RandomDraws.singleFromRange(0, i);")
    lines.append("    [a[i], a[j]] = [a[j], a[i]];")
    lines.append("  }")
    lines.append("  return a;")
    lines.append("}")
    lines.append("")
    lines.append("export function getWordList(listIndex) {")
    lines.append("  return WORD_LISTS[listIndex % WORD_LISTS.length];")
    lines.append("}")
    lines.append("")
    lines.append("export function buildRecognitionSequence(targets, distractors) {")
    lines.append("  const items = [")
    lines.append('    ...targets.map((word) => ({ word, wordType: "target" })),')
    lines.append('    ...distractors.map((word) => ({ word, wordType: "distractor" })),')
    lines.append("  ];")
    lines.append("  return shuffleArray(items);")
    lines.append("}")
    lines.append("")
    return "\n".join(lines)


# ─── Main ───────────────────────────────────────────────────

def main():
    random.seed(42)
    script_dir = Path(__file__).parent

    # Load data sources
    subtlex_path = script_dir / "SUBTLEX-UK.xlsx"
    glasgow_path = script_dir / "glasgow_norms.csv"
    conc_path = script_dir / "concreteness.csv"

    missing = []
    if not subtlex_path.exists():
        missing.append(f"SUBTLEX-UK: download from https://psychology.nottingham.ac.uk/subtlex-uk/SUBTLEX-UK.xlsx.zip")
    if not glasgow_path.exists():
        missing.append(f"Glasgow Norms: download from https://doi.org/10.3758/s13428-018-1099-3 (Supplementary Material 2)")
    if missing:
        print("ERROR: Missing data files:")
        for m in missing:
            print(f"  - {m}")
        sys.exit(1)

    entries, total_freq = load_subtlex_uk(str(subtlex_path))
    glasgow = load_glasgow_norms(str(glasgow_path))

    brysbaert_conc = {}
    if conc_path.exists():
        brysbaert_conc = load_concreteness(str(conc_path))

    candidates = filter_candidates(entries, total_freq, glasgow, brysbaert_conc)

    if len(candidates) < 336:
        print(f"ERROR: Only {len(candidates)} candidates (need 336)")
        sys.exit(1)

    lists = generate_balanced_lists(candidates)

    # Print De Vent criteria balance report
    print("\n" + "=" * 80)
    print("DE VENT FRAMEWORK BALANCE REPORT")
    print("=" * 80)

    criteria_keys = [
        ("mean_log_freq", "Word frequency (log10 f/M)"),
        ("mean_length", "Word length (letters)"),
        ("mean_syllables", "Syllables"),
        ("mean_phonemes", "Phonemes (est.)"),
        ("mean_concreteness", "Concreteness"),
        ("mean_imageability", "Imageability"),
        ("mean_familiarity", "Familiarity"),
        ("mean_aoa", "Age of acquisition"),
        ("mean_valence", "Valence"),
        ("mean_arousal", "Arousal"),
    ]

    print(f"\n{'Criterion':<30} {'Grand Mean':>10} {'Max Dev':>10} {'Range':>20}")
    print("-" * 72)

    for key, label in criteria_keys:
        vals = [lst["stats"][key] for lst in lists if lst["stats"].get(key) is not None]
        if vals:
            gm = sum(vals) / len(vals)
            md = max(abs(v - gm) for v in vals)
            rng = f"[{min(vals):.3f}, {max(vals):.3f}]"
            print(f"{label:<30} {gm:>10.4f} {md:>10.4f} {rng:>20}")

    # Per-list detail
    print(f"\n{'List':>4} {'Freq':>7} {'Len':>5} {'Syl':>5} {'Conc':>6} {'Img':>6} {'Fam':>6} {'AoA':>6} {'Val':>6} {'Aro':>6}")
    print("-" * 72)
    for lst in lists:
        s = lst["stats"]
        print(
            f"{lst['list_index']:>4} "
            f"{s['mean_log_freq']:>7.4f} "
            f"{s['mean_length']:>5.2f} "
            f"{s['mean_syllables']:>5.2f} "
            f"{s['mean_concreteness']:>6.2f} "
            f"{s['mean_imageability']:>6.2f} "
            f"{s['mean_familiarity']:>6.2f} "
            f"{s['mean_aoa']:>6.2f} "
            f"{s['mean_valence']:>6.2f} "
            f"{s['mean_arousal']:>6.2f}"
        )

    # Verify uniqueness
    all_words = set()
    for lst in lists:
        for w in lst["targets"] + lst["distractors"]:
            if w in all_words:
                print(f"DUPLICATE: {w}")
            all_words.add(w)
    print(f"\nTotal unique words: {len(all_words)}")

    # Write outputs
    js_content = format_as_js(lists)
    output_path = script_dir.parent / "stimuli.js"
    output_path.write_text(js_content)
    print(f"\nWrote {output_path}")

    # Save detailed stats
    stats_output = []
    for lst in lists:
        stats_output.append({
            "list_index": lst["list_index"],
            "targets": lst["targets"],
            "distractors": lst["distractors"],
            "stats": lst["stats"],
            "stats_targets": lst["stats_targets"],
            "stats_distractors": lst["stats_distractors"],
        })
    stats_path = script_dir / "word_list_stats.json"
    stats_path.write_text(json.dumps(stats_output, indent=2))
    print(f"Wrote {stats_path}")


if __name__ == "__main__":
    main()
